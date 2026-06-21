# -*- coding: utf-8 -*-
"""Build data/foodland_knowledge.json from the 7 xlsx knowledge sources plus
the directly-supplied foodland_recipe_ingredients.json.

No invented data: every URL/title/price comes from a real cell (hyperlink
target preferred, displayed value as fallback) or from data/products.json
(the real feed). Missing translations/links are stored as "", never guessed.
Idempotent: re-running over the same inputs always produces the same JSON.
"""
import json
import re
from pathlib import Path

import openpyxl

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR.parent / "data"
OUT_PATH = DATA_DIR / "foodland_knowledge.json"

LANGS = ["SK", "CZ", "AT", "EN", "PL", "HU", "VI"]

# en dash (U+2013) separates "{title} – {reason}" in suggestion cells.
# Regular hyphens can appear inside titles themselves, so split on en dash only.
EN_DASH = "–"


def cell_url(cell):
    """Real URL for a language cell: hyperlink target, else cell value if it
    looks like a URL, else "" (covers None and the literal '-'/'–' markers)."""
    if cell.hyperlink is not None and cell.hyperlink.target:
        return cell.hyperlink.target
    v = cell.value
    if isinstance(v, str) and v.strip().startswith("http"):
        return v.strip()
    return ""


def cell_text(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def lang_urls(ws, row, start_col):
    """Read LANGS in order starting at start_col, one column per language."""
    return {lang: cell_url(ws.cell(row=row, column=start_col + i)) for i, lang in enumerate(LANGS)}


def split_suggestion(cell):
    """Parse a cross-sell/alternative suggestion cell:
    value = '{title} – {reason}' (en dash separated), hyperlink = product link.
    Returns None if the cell is empty."""
    if cell.value is None:
        return None
    text = str(cell.value)
    link = cell.hyperlink.target if cell.hyperlink else ""
    if EN_DASH in text:
        title, reason = text.split(EN_DASH, 1)
        title = title.strip()
        reason = reason.strip()
    else:
        title, reason = text.strip(), ""
    return {"title": title, "link": link, "reason": reason}


def load_faq():
    wb = openpyxl.load_workbook(DATA_DIR / "foodland_faq_tabulka.xlsx", data_only=False)
    ws = wb["FAQ"]
    out = []
    for r in range(2, ws.max_row + 1):
        category = cell_text(ws.cell(row=r, column=1))
        question = cell_text(ws.cell(row=r, column=2))
        if not question:
            continue
        out.append({
            "category": category,
            "question": question,
            "answer": cell_text(ws.cell(row=r, column=3)),
            "urls": lang_urls(ws, r, 4),
            "note": cell_text(ws.cell(row=r, column=11)),
        })
    return out


def load_recipes_base():
    wb = openpyxl.load_workbook(DATA_DIR / "foodland_recepty_jazykove_mutacie.xlsx", data_only=False)
    ws = wb["Recepty Foodland"]
    out = []
    for r in range(2, ws.max_row + 1):
        name = cell_text(ws.cell(row=r, column=2))
        if not name:
            continue
        out.append({
            "cuisine": cell_text(ws.cell(row=r, column=1)),
            "name": name,
            "urls": lang_urls(ws, r, 3),
            "note": cell_text(ws.cell(row=r, column=10)),
        })
    return out


def load_blog():
    wb = openpyxl.load_workbook(DATA_DIR / "foodland_magazin_clanky_jazykove_mutacie.xlsx", data_only=False)
    ws = wb["Magazín články"]
    out = []
    for r in range(2, ws.max_row + 1):
        title = cell_text(ws.cell(row=r, column=2))
        if not title:
            continue
        out.append({
            "theme": cell_text(ws.cell(row=r, column=1)),
            "title": title,
            "urls": lang_urls(ws, r, 3),
            "note": cell_text(ws.cell(row=r, column=10)),
        })
    return out


def load_suggestion_table(filename, sheet, n_suggestions):
    wb = openpyxl.load_workbook(DATA_DIR / filename, data_only=False)
    ws = wb[sheet]
    out = []
    for r in range(2, ws.max_row + 1):
        pid = cell_text(ws.cell(row=r, column=1))
        if not pid:
            continue
        product_cell = ws.cell(row=r, column=2)
        suggestions = []
        for i in range(n_suggestions):
            s = split_suggestion(ws.cell(row=r, column=5 + i))
            if s is not None:
                suggestions.append(s)
        out.append({
            "product_id": pid,
            "product_title": cell_text(product_cell),
            "product_link": product_cell.hyperlink.target if product_cell.hyperlink else "",
            "price": cell_text(ws.cell(row=r, column=3)),
            "category": cell_text(ws.cell(row=r, column=4)),
            "suggestions": suggestions,
        })
    return out


def load_products_ai(products_by_id):
    wb = openpyxl.load_workbook(DATA_DIR / "foodland_products_ai_tabulka.xlsx", data_only=False)
    ws = wb["Products_AI"]
    out = []
    for r in range(2, ws.max_row + 1):
        pid = cell_text(ws.cell(row=r, column=1))
        if not pid:
            continue
        feed_product = products_by_id.get(pid)
        # Title/link/price for this row come from the real feed by ID --
        # the xlsx's own "Produkt (URL)" column only carries the link.
        url_cell = ws.cell(row=r, column=2)
        link = url_cell.hyperlink.target if url_cell.hyperlink else cell_text(url_cell)
        title = feed_product["title"] if feed_product else ""
        price = feed_product["price"] if feed_product else cell_text(ws.cell(row=r, column=5))

        attrs_raw = cell_text(ws.cell(row=r, column=8))
        diet_tags = [a.strip() for a in attrs_raw.split(",") if a.strip() and a.strip() != "-"]

        ai_description = {}
        for i, lang in enumerate(LANGS):
            ai_description[lang] = cell_text(ws.cell(row=r, column=9 + i))

        cross_sell = split_suggestion(ws.cell(row=r, column=16))
        alternative = split_suggestion(ws.cell(row=r, column=17))

        recipe_cell = ws.cell(row=r, column=18)
        related_recipe = None
        if recipe_cell.value is not None:
            related_recipe = {
                "name": cell_text(recipe_cell),
                "link": recipe_cell.hyperlink.target if recipe_cell.hyperlink else "",
            }

        out.append({
            "id": pid,
            "title": title,
            "link": link,
            "category": cell_text(ws.cell(row=r, column=3)),
            "cuisine": cell_text(ws.cell(row=r, column=4)),
            "price": price,
            "package_size": cell_text(ws.cell(row=r, column=6)),
            "gtin": cell_text(ws.cell(row=r, column=7)),
            "diet_tags": diet_tags,
            "ai_description": ai_description,
            "cross_sell": cross_sell,
            "alternative": alternative,
            "related_recipe": related_recipe,
            "note": cell_text(ws.cell(row=r, column=19)),
        })
    return out


def load_intent_mapping():
    wb = openpyxl.load_workbook(DATA_DIR / "foodland_intentmapping_tabulka.xlsx", data_only=False)
    ws = wb["IntentMapping"]
    out = []
    for r in range(2, ws.max_row + 1):
        intent = cell_text(ws.cell(row=r, column=2))
        if not intent:
            continue
        out.append({
            "type": cell_text(ws.cell(row=r, column=1)),
            "intent": intent,
            "urls": lang_urls(ws, r, 3),
            "content_name": cell_text(ws.cell(row=r, column=10)),
            "note": cell_text(ws.cell(row=r, column=11)),
        })
    return out


def merge_recipe_ingredients(recipes_base):
    """Merge foodland_recipe_ingredients.json into recipes_base by exact
    name match. Never invents ingredients for an unmatched recipe."""
    ing_path = DATA_DIR / "foodland_recipe_ingredients.json"
    ing_by_name = {}
    if ing_path.exists():
        with open(ing_path, encoding="utf-8") as f:
            for rec in json.load(f):
                ing_by_name[rec["name"]] = rec

    merged = []
    for base in recipes_base:
        rec = dict(base)
        ing = ing_by_name.get(base["name"])
        if ing is None:
            rec["ingredients_available"] = False
            rec["ingredient_source_lang"] = ""
            rec["ingredient_source_url"] = ""
            rec["page_template"] = ""
            rec["ingredients"] = []
            rec["curated_shop_links"] = []
        else:
            has_ingredients = len(ing.get("ingredients", [])) > 0
            rec["ingredients_available"] = has_ingredients
            rec["ingredient_source_lang"] = ing.get("ingredient_source_lang", "")
            rec["ingredient_source_url"] = ing.get("ingredient_source_url", "")
            rec["page_template"] = ing.get("page_template", "")
            rec["ingredients"] = ing.get("ingredients", [])
            rec["curated_shop_links"] = ing.get("curated_shop_links", [])
        merged.append(rec)
    return merged


def main():
    products = json.load(open(DATA_DIR / "products.json", encoding="utf-8"))
    products_by_id = {p["id"]: p for p in products}

    recipes_base = load_recipes_base()
    recipes = merge_recipe_ingredients(recipes_base)

    knowledge = {
        "faq": load_faq(),
        "recipes": recipes,
        "blog": load_blog(),
        "cross_sell": load_suggestion_table("foodland_crosssell_tabulka.xlsx", "Cross-sell", 5),
        "alternatives": load_suggestion_table("foodland_alternativy_tabulka.xlsx", "Alternativy", 5),
        "products_ai": load_products_ai(products_by_id),
        "intent_mapping": load_intent_mapping(),
    }

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(knowledge, f, ensure_ascii=False, indent=2)

    print(f"wrote knowledge to {OUT_PATH}")
    for k, v in knowledge.items():
        print(f"  {k}: {len(v)}")


if __name__ == "__main__":
    main()
